import re
import pandas as pd
import win32com.client
import pywintypes
import os
import json
import math
import time
import shutil
import tkinter as tk
from tkinter import messagebox,filedialog,simpledialog
from tkinter import ttk
from collections import deque
from openpyxl import load_workbook
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP
from copy import copy
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter


DEBUG = True
output_path = os.path.join(os.getcwd(), "name_wh_output.xlsx")
SMALL_FACE_COL = "H"
PRODUCTION_ROW_HEIGHT = 24.95

BASE_DIR = os.path.dirname(os.path.dirname(__file__))
BACKUP_DIR = os.path.join(BASE_DIR, "backup")
PROGRESS_JSON_PATH = os.path.join(BACKUP_DIR, "progress.json")
PROGRESS_HISTORY_DIR = os.path.join(BACKUP_DIR, "history")
LAST_SUCCESS_JSON_PATH = os.path.join(BACKUP_DIR, "last_success.json")

class DiscardCurrentGroup(Exception):
    pass


class EndWholeInput(Exception):
    """结束整个录入流程"""
    pass

class FloatingPanel:
    def __init__(self):
        self.root = tk.Tk()
        self.root.withdraw()
        self.root.title("CAD 录入助手")

        self.root.attributes("-topmost", True)

        self.root.protocol("WM_DELETE_WINDOW", self.on_close_requested)

        self.root.resizable(False, False)

        # 状态变量
        self.step_var = tk.StringVar(value="等待开始")
        self.mode_var = tk.StringVar(value="未确定")
        self.code_var = tk.StringVar(value="-")
        self.length_var = tk.StringVar(value="-")
        self.width_var = tk.StringVar(value="-")

        self.recent_records = deque(maxlen=5)
        self.current_message = "暂无消息"

        # UI
        title = tk.Label(
            self.root, text="CAD 录入助手",
            font=("Microsoft JhengHei UI", 12, "bold")
        )
        title.pack(pady=(10, 8))

        info_frame = tk.Frame(self.root)
        info_frame.pack(fill="x", padx=10)

        tk.Label(info_frame, text="当前步骤：", anchor="w").grid(row=0, column=0, sticky="w")
        tk.Label(info_frame, textvariable=self.step_var, anchor="w", fg="blue").grid(row=0, column=1, sticky="w")

        tk.Label(info_frame, text="当前模式：", anchor="w").grid(row=1, column=0, sticky="w")
        tk.Label(info_frame, textvariable=self.mode_var, anchor="w", fg="purple").grid(row=1, column=1, sticky="w")

        temp_frame = tk.LabelFrame(self.root, text="当前临时数据")
        temp_frame.pack(fill="x", padx=10, pady=10)

        tk.Label(temp_frame, text="编号：", anchor="w").grid(row=0, column=0, sticky="w", padx=8, pady=4)
        tk.Label(temp_frame, textvariable=self.code_var, anchor="w").grid(row=0, column=1, sticky="w", padx=8, pady=4)

        tk.Label(temp_frame, text="长：", anchor="w").grid(row=1, column=0, sticky="w", padx=8, pady=4)
        tk.Label(temp_frame, textvariable=self.length_var, anchor="w").grid(row=1, column=1, sticky="w", padx=8, pady=4)

        tk.Label(temp_frame, text="宽：", anchor="w").grid(row=2, column=0, sticky="w", padx=8, pady=4)
        tk.Label(temp_frame, textvariable=self.width_var, anchor="w").grid(row=2, column=1, sticky="w", padx=8, pady=4)

        message_frame = tk.LabelFrame(self.root, text="状态提示")
        message_frame.pack(fill="both", padx=10, pady=(0, 10))

        message_inner = tk.Frame(message_frame)
        message_inner.pack(fill="both", expand=True, padx=5, pady=5)

        self.message_text = tk.Text(
            message_inner,
            height=2,
            width=46,
            font=("Microsoft JhengHei UI", 9),
            wrap="word"
        )
        self.message_text.pack(side="left", fill="both", expand=True)

        message_scrollbar = tk.Scrollbar(message_inner, command=self.message_text.yview)
        message_scrollbar.pack(side="right", fill="y")

        self.message_text.config(yscrollcommand=message_scrollbar.set, state="disabled")

        recent_frame = tk.LabelFrame(self.root, text="最近记录（最多5组）")
        recent_frame.pack(fill="both", expand=True, padx=10, pady=(0, 10))

        self.recent_text = tk.Text(
            recent_frame, height=10, width=40,
            font=("Consolas", 10)
        )
        self.recent_text.pack(fill="both", expand=True, padx=5, pady=5)
        self.recent_text.config(state="disabled")

        self.place_top_right()
        self.root.deiconify() 
        self.root.lift()
        self.root.focus_force()
        self.refresh()

    def on_close_requested(self):
        messagebox.showwarning(
            "提示",
            "这个状态浮窗不能直接关闭。\n请在录入流程里按取消/结束退出。",
            parent=self.root
        )

    def refresh(self):
        self.root.update_idletasks()
        self.root.update()

    def set_step(self, text):
        self.step_var.set(text)
        self.refresh()

    def set_mode(self, text):
        self.mode_var.set(text)
        self.refresh()

    def set_temp(self, code=None, length=None, width=None):
        if code is not None:
            self.code_var.set(str(code))
        if length is not None:
            self.length_var.set(str(length))
        if width is not None:
            self.width_var.set(str(width))
        self.refresh()

    def clear_temp(self):
        self.code_var.set("-")
        self.length_var.set("-")
        self.width_var.set("-")
        self.refresh()

    def set_status(self, text):
        self.current_message = str(text)
        self._render_messages()
        self.refresh()

    def clear_status(self):
        self.current_message = "暂无消息"
        self._render_messages()
        self.refresh()
        
    def _render_messages(self):
        self.message_text.config(state="normal")
        self.message_text.delete("1.0", tk.END)
        self.message_text.insert(tk.END, self.current_message)
        self.message_text.config(state="disabled")
    
    def add_record(self, record):
        """
        record 例子：
        {'编号': 'AR-1', '长': 597.62, '宽': 243.10}
        """
        self.recent_records.appendleft(record)
        self._render_recent()
        self.refresh()

    def place_top_right(self, width=440, height=460, margin_x=20, margin_y=40):
        self.root.update_idletasks()

        screen_w = self.root.winfo_screenwidth()
        screen_h = self.root.winfo_screenheight()

        x = max(0, screen_w - width - margin_x)
        y = max(0, margin_y)

        self.root.geometry(f"{width}x{height}+{x}+{y}")

    def _render_recent(self):
        self.recent_text.config(state="normal")
        self.recent_text.delete("1.0", tk.END)

        if not self.recent_records:
            self.recent_text.insert(tk.END, "暂无记录")
        else:
            for i, r in enumerate(self.recent_records, start=1):
                line1 = f"{i}. {r['编号']} | 长:{r['长']} | 宽:{r['宽']}\n"

                line2_parts = []
                if r.get("长边1加工"):
                    line2_parts.append(f"长1:{r['长边1加工']}")
                if r.get("长边2加工"):
                    line2_parts.append(f"长2:{r['长边2加工']}")
                if r.get("短边1加工"):
                    line2_parts.append(f"短1:{r['短边1加工']}")
                if r.get("短边2加工"):
                    line2_parts.append(f"短2:{r['短边2加工']}")
                if r.get("左上角加工"):
                    line2_parts.append(f"左上:{r['左上角加工']}")
                if r.get("右上角加工"):
                    line2_parts.append(f"右上:{r['右上角加工']}")
                if r.get("左下角加工"):
                    line2_parts.append(f"左下:{r['左下角加工']}")
                if r.get("右下角加工"):
                    line2_parts.append(f"右下:{r['右下角加工']}")

                if line2_parts:
                    line2 = "    加工：" + " | ".join(line2_parts) + "\n"
                else:
                    line2 = "    加工：无\n"

                self.recent_text.insert(tk.END, line1)
                self.recent_text.insert(tk.END, line2)
                self.recent_text.insert(tk.END, "\n")

        self.recent_text.config(state="disabled")

    def destroy(self):
        self.root.destroy()

# 与 CAD 交互的函数
def get_autocad_document():
    acad = win32com.client.Dispatch("AutoCAD.Application")
    doc = acad.ActiveDocument
    return acad, doc

# 内容清洗
def clean_cad_text(text: str) -> str:
    if text is None:
        return ""

    text = str(text).strip()

    # CAD 多行文字换行 / 标注上下分隔
    text = text.replace("\\P", "")
    text = text.replace("\\X", "")

    # 去掉大括号包住的 MTEXT 格式头
    # 例如：{\fSimSun|b0|i0|c134|p2;A1-4}
    text = re.sub(r"\{\\[^;]*;", "", text)

    text = re.sub(r"\\[A-Za-z][^;]*;", "", text)
    text = text.replace("{", "").replace("}", "")

    return text.strip()

# 选择对象函数
def get_entity_by_click(doc, prompt="请选择对象", parent_window=None, allow_finish_by_empty=False):


    fail_count = 0

    while True:
        print(prompt)
        time.sleep(0.25)

        try:
            obj, point = doc.Utility.GetEntity()
            return obj

        except pywintypes.com_error as e:
            fail_count += 1

            if DEBUG:
                print(f"[调试] GetEntity com_error: {repr(e)}")

            if allow_finish_by_empty and fail_count >= 2:
                return None

            if fail_count < 2:
                print("没有选到有效对象，请再点一次。")
                continue

            action = ask_fail_action(
                parent_window=parent_window,
                message="连续多次没有选到有效对象。"
            )

            if action == "retry":
                fail_count = 0
                print("继续当前步骤，请重新选择。")
                continue
            elif action == "discard_group":
                raise DiscardCurrentGroup()
            elif action == "end_all":
                raise EndWholeInput()

        except Exception as e:
            fail_count += 1
            print(f"选择对象失败：{e}")

            if fail_count < 2:
                print("请再点一次。")
                continue

            action = ask_fail_action(
                parent_window=parent_window,
                message=f"连续多次选择失败。\n\n错误信息：{e}"
            )

            if action == "retry":
                fail_count = 0
                print("继续当前步骤，请重新选择。")
                continue
            elif action == "discard_group":
                raise DiscardCurrentGroup()
            elif action == "end_all":
                raise EndWholeInput()
            
# 读取对象属性的函数，增加了重试机制应对偶尔的 COM 访问问题
def safe_get_object_name(entity, retry=2, delay=0.08):
    for i in range(retry + 1):
        try:
            return entity.ObjectName
        except Exception as e:
            if i < retry:
                time.sleep(delay)
            else:
                raise e

def read_code_entity(entity):
    if entity is None:
        return None,"读取编号失败：entity is None"

    try:
        obj_name = safe_get_object_name(entity)
    except Exception as e:
        return None, f"读取编号失败：无法获取 ObjectName,{e}"

    if obj_name not in ["AcDbText", "AcDbMText"]:
        return None, f"选中的不是文字对象，而是：{obj_name}"

    try:
        raw_text = entity.TextString
        text = clean_cad_text(raw_text)
    except Exception as e:
        return None, f"读取编号失败：{e}"

    if DEBUG:
        print(f"[编号] raw={repr(raw_text)} -> clean={repr(text)}")

    return text, None

def extract_number_from_text(text):
    if text is None:
        return None

    text = clean_cad_text(text)
    text = str(text).strip()

    if not text:
        return None

    match = re.search(r'-?\d+(?:\.\d+)?', text)
    if not match:
        return None

    try:
        return float(match.group())
    except Exception:
        return None


def read_size_entity(entity):
    if entity is None:
        return None, "读取尺寸失败：entity is None"

    try:
        obj_name = safe_get_object_name(entity)
    except Exception as e:
        return None, f"读取尺寸失败：无法获取 ObjectName，{e}"

    if DEBUG:
        print(f"[调试] read_size_entity -> obj_name={obj_name}")

    if "Dimension" in obj_name:
        return read_dimension_entity(entity)

    if obj_name in ["AcDbText", "AcDbMText"]:
        try:
            raw_text = entity.TextString
            clean_text = clean_cad_text(raw_text)

            if DEBUG:
                print(f"[尺寸文字] raw={repr(raw_text)} -> clean={repr(clean_text)}")

            value = extract_number_from_text(clean_text)
            if value is None:
                return None, f"选中的文字不是可识别尺寸：{clean_text}"

            value = cad_round_1(value)
            return value, None

        except Exception as e:
            return None, f"读取文字尺寸失败：{e}"

    return None, f"当前步骤只允许选择：标注对象 或 文字对象。你选的是：{obj_name}"

def cad_round_1(value):
    if value in [None, ""]:
        return value

    d = Decimal(repr(float(value))) + Decimal("0.000001")
    return float(d.quantize(Decimal("0.1"), rounding=ROUND_HALF_UP))


def cad_round_0(value):
    if value in [None, ""]:
        return value

    d = Decimal(repr(float(value))) + Decimal("0.000001")
    return int(d.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def read_dimension_entity(entity):
    if entity is None:
        return None, "读取标注失败：entity is None"

    try:
        obj_name = safe_get_object_name(entity)
    except Exception as e:
        return None, f"读取标注失败：无法获取 ObjectName，{e}"

    if DEBUG:
        print(f"[调试] read_dimension_entity -> obj_name={obj_name}")

    if "Dimension" not in obj_name:
        return None, f"选中的不是标注对象，而是：{obj_name}"

    try:
        txt = str(getattr(entity, "TextOverride", "") or "").strip()
        measurement = float(entity.Measurement)

        if DEBUG:
            print(f"[标注调试] TextOverride={repr(txt)}")
            print(f"[标注调试] Measurement={measurement}")

        match = re.search(r';\s*(-?\d+(?:\.\d+)?)\s*}?$', txt)
        if match:
            value = cad_round_1(match.group(1))
            if DEBUG:
                print(f"[标注] 使用显示值(TextOverride分号后明确数字): {value}")
            return value, None

        if "<>" in txt:
            value = float(measurement)
            if DEBUG:
                print(f"[标注] TextOverride含<>，改用Measurement原始值: {value}")
            return value, None

        pure_txt = re.sub(r"{.*?;|[{}\\\\]", "", txt).strip()
        if re.fullmatch(r'-?\d+(?:\.\d+)?', pure_txt):
            value = cad_round_1(pure_txt)
            if DEBUG:
                print(f"[标注] 使用纯数字TextOverride: {value}")
            return value, None

        value = float(measurement)
        if DEBUG:
            print(f"[标注] 使用测量值(Measurement原始值): {value}")
        return value, None

    except Exception as e:
        return None, f"读取标注失败：对象是 {obj_name}，错误：{e}"

def export_to_excel(data, output_path="output.xlsx"):
    base, ext = os.path.splitext(output_path)
    i = 1
    new_path = output_path

    while os.path.exists(new_path):
        new_path = f"{base}_{i}{ext}"
        i += 1

    df = pd.DataFrame(data)
    df.to_excel(new_path, index=False)
    return df, new_path

def get_entity_type(entity):
    if entity is None:
        return None

    try:
        obj_name = safe_get_object_name(entity)
    except Exception:
        return None

    if obj_name in ["AcDbText", "AcDbMText"]:
        return "text"

    if "Dimension" in obj_name:
        return "dim"

    return None

def get_entity_type_with_name(entity):
    if entity is None:
        return None, "对象为空"

    try:
        obj_name = safe_get_object_name(entity)
    except Exception as e:
        return None, f"无法获取对象类型：{e}"

    if obj_name in ["AcDbText", "AcDbMText"]:
        return "text", obj_name

    if "Dimension" in obj_name:
        return "dim", obj_name

    return None, obj_name

def center_window(parent, win, width, height):
    win.update_idletasks()

    screen_w = win.winfo_screenwidth()
    screen_h = win.winfo_screenheight()

    x = (screen_w - width) // 2
    y = (screen_h - height) // 2

    # 防越界
    x = max(0, min(x, screen_w - width))
    y = max(0, min(y, screen_h - height))

    win.geometry(f"{width}x{height}+{x}+{y}")

def ask_fail_action(parent_window=None, message="连续多次没有选到有效对象。"):
    win = tk.Toplevel(parent_window)
    win.title("操作选择")

    result = {"action": "retry"}

    tk.Label(
        win,
        text=message,
        justify="left",
        anchor="w",
        font=("Microsoft JhengHei UI", 10)
    ).pack(fill="x", padx=20, pady=(18, 10))

    tk.Label(
        win,
        text="请选择接下来要怎么处理：",
        justify="left",
        anchor="w"
    ).pack(fill="x", padx=20, pady=(0, 12))

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=(0, 16))

    def choose_retry():
        result["action"] = "retry"
        win.destroy()

    def choose_discard():
        result["action"] = "discard_group"
        win.destroy()

    def choose_end():
        result["action"] = "end_all"
        win.destroy()

    tk.Button(btn_frame, text="继续当前步骤", width=12, command=choose_retry).pack(side="left", padx=6)
    tk.Button(btn_frame, text="放弃当前组", width=12, command=choose_discard).pack(side="left", padx=6)
    tk.Button(btn_frame, text="结束录入", width=12, command=choose_end).pack(side="left", padx=6)

    prepare_dialog(win, parent_window, 420, 170)
    win.wait_window()
    return result["action"]

def is_code_text(text):
    if not text:
        return False

    text = clean_cad_text(str(text).strip())

    if re.fullmatch(r'-?\d+(?:\.\d+)?(?:\s*(mm|cm|MM|CM))?', text):
        return False
    
    has_digit = any(c.isdigit() for c in text)
    has_dash = "-" in text

    return has_digit and has_dash

def prepare_dialog(win, parent_window, width, height):
    win.withdraw()

    win.resizable(False, False)

    if parent_window is not None and parent_window.winfo_exists():
        win.transient(parent_window)

    win.update_idletasks()

    center_window(parent_window, win, width, height)

    win.deiconify()

    if parent_window is not None and parent_window.winfo_exists():
        parent_window.lift()

    win.lift()
    win.attributes("-topmost", True)

    try:
        win.focus_force()
    except Exception:
        pass

    win.grab_set()

# 根据选择的对象分类，增加对无效对象的处理
def classify_post_width_entity(entity):
    if entity is None:
        return "invalid", "对象为空"

    ent_type, obj_name = get_entity_type_with_name(entity)

    # 文字 / 多行文字：进一步判断是不是编号
    if ent_type == "text":
        text, err = read_code_entity(entity)
        if text is None:
            return "invalid", err

        if is_code_text(text):
            return "next_code", text

        return "non_code", {
            "object_type": obj_name,
            "display_name": text
        }

    try:
        display_name = getattr(entity, "EffectiveName", None)
        if not display_name:
            display_name = getattr(entity, "Name", None)
        if not display_name and ent_type == "text":
            display_name = getattr(entity, "TextString", None)
        if not display_name:
            display_name = obj_name
    except Exception:
        display_name = obj_name

    try:
        handle = getattr(entity, "Handle", "")
    except Exception:
        handle = ""

    try:
        insertion_point = getattr(entity, "InsertionPoint", None)
    except Exception:
        insertion_point = None

    return "non_code", {
        "object_type": obj_name,
        "display_name": str(display_name),
        "handle": str(handle),
        "insertion_point": insertion_point
    }
    
def is_duplicate_small_face(candidate, small_face_candidates):
    handle = str(candidate.get("handle", "")).strip()
    if not handle:
        return False

    for item in small_face_candidates:
        if str(item.get("handle", "")).strip() == handle:
            return True

    return False

def split_code(code):
    return str(code).strip().split("-")

def merge_range_part(part1, part2):
    part1 = str(part1).strip()
    part2 = str(part2).strip()

    if part1.isdigit() and part2.isdigit():
        a, b = int(part1), int(part2)
        return f"{min(a, b)}~{max(a, b)}"

    if part1.isalpha() and part2.isalpha():
        a, b = sorted([part1, part2])
        return f"{a}~{b}"

    m1 = re.fullmatch(r"(\d+)([A-Za-z]+)", part1)
    m2 = re.fullmatch(r"(\d+)([A-Za-z]+)", part2)
    if m1 and m2:
        num1, suffix1 = int(m1.group(1)), m1.group(2)
        num2, suffix2 = int(m2.group(1)), m2.group(2)

        if suffix1 == suffix2:
            start, end = sorted([num1, num2])
            return f"{start}{suffix1}~{end}{suffix1}"

    return f"{part1}~{part2}"


def merge_code_range(code1, code2):
    parts1 = split_code(code1)
    parts2 = split_code(code2)

    # 结构不同，不合并
    if len(parts1) != len(parts2):
        print("编号结构不同，不能合并")
        return None

    diff_indices = [i for i, (a, b) in enumerate(zip(parts1, parts2)) if a != b]

    if len(diff_indices) == 0:
        return code1

    if len(diff_indices) > 1:
        print("编号有多个位置不同，不能合并")
        return None

    diff_i = diff_indices[0]

    merged_parts = parts1.copy()
    merged_parts[diff_i] = merge_range_part(parts1[diff_i], parts2[diff_i])

    return "-".join(merged_parts)

def pick_dimension_with_retry(doc, panel, step_text, prompt):
    panel.set_step(step_text)

    while True:
        obj, ent_type, err = pick_entity_with_retry(
            doc, panel,
            step_text=step_text,
            prompt=prompt,
            allowed_types=["text", "dim"]
        )
        if err == "cancel":
            return None, "cancel"

        value, err_msg = read_size_entity(obj)
        if value is None:
            print(err_msg)
            panel.set_status(err_msg)
            continue

        panel.clear_status()
        return value, None

def pick_code_with_retry(doc, panel, step_text, prompt):
    panel.set_step(step_text)

    while True:
        obj, ent_type, err = pick_entity_with_retry(
            doc, panel,
            step_text=step_text,
            prompt=prompt,
            allowed_types=["text"]
        )
        if err == "cancel":
            return None, "cancel"

        value, err_msg = read_code_entity(obj)
        if value is None:
            print(err_msg)
            panel.set_status(err_msg)
            continue

        panel.clear_status()
        return value, None
def pick_second_entity_with_retry(doc, panel):
    return pick_entity_with_retry(
        doc, panel,
        step_text="请选择（编号 或 长尺寸）",
        prompt="请选择编号或长尺寸（标注或文字）",
        allowed_types=["text", "dim"]
    )

def pick_entity_with_retry(doc, panel, step_text, prompt, allowed_types):
    panel.set_step(step_text)

    while True:
        print(f"下一步：{prompt}")

        try:
            obj = get_entity_by_click(doc, prompt, panel.root)
        except DiscardCurrentGroup:
            panel.set_status("已放弃当前组，请重新录入。")
            panel.clear_temp()
            raise
        except EndWholeInput:
            msg = "用户结束录入。"
            print(msg)
            panel.set_status(msg)
            return None, None, "cancel"

        ent_type, obj_name = get_entity_type_with_name(obj)

        if ent_type in allowed_types:
            panel.clear_status()
            return obj, ent_type, None

        allow_map = {
            "text": "文字对象",
            "dim": "标注对象"
        }
        allow_text = " / ".join(allow_map[t] for t in allowed_types)

        msg = f"当前步骤只允许选择：{allow_text}。你选的是：{obj_name}"
        print(msg)
        panel.set_status(msg)
        
def find_total_row(ws, start_row, max_scan=300):
    for r in range(start_row, start_row + max_scan):
        value = ws.cell(r, 1).value
        if str(value).strip() == "合计":
            return r
    return None


def copy_row_style_and_values(ws, src_row, dst_row, min_col=1, max_col=10):
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height

    for col in range(min_col, max_col + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)

        if src.has_style:
            dst._style = copy(src._style)

        if src.number_format:
            dst.number_format = src.number_format

        if src.font:
            dst.font = copy(src.font)

        if src.fill:
            dst.fill = copy(src.fill)

        if src.border:
            dst.border = copy(src.border)

        if src.alignment:
            dst.alignment = copy(src.alignment)

        if src.protection:
            dst.protection = copy(src.protection)

        if isinstance(src.value, str) and src.value.startswith("="):
            src_addr = f"{get_column_letter(col)}{src_row}"
            dst_addr = f"{get_column_letter(col)}{dst_row}"
            try:
                dst.value = Translator(src.value, origin=src_addr).translate_formula(dst_addr)
            except Exception:
                dst.value = src.value
        else:
            dst.value = src.value


def ensure_template_data_rows(ws, start_row, data_count, min_col=1, max_col=10):
    total_row = find_total_row(ws, start_row)

    if total_row is None:
        return

    existing_data_rows = total_row - start_row
    need_extra_rows = data_count - existing_data_rows

    if need_extra_rows <= 0:
        return

    template_row = total_row - 1

    for _ in range(need_extra_rows):
        ws.insert_rows(total_row)

        copy_row_style_and_values(
            ws,
            src_row=template_row,
            dst_row=total_row,
            min_col=min_col,
            max_col=max_col
        )

        total_row += 1
        template_row += 1
        
def apply_production_row_height(ws, start_row, data_count, row_height=PRODUCTION_ROW_HEIGHT, include_total=True):
    if data_count <= 0:
        return

    end_row = start_row + data_count - 1

    for row_num in range(start_row, end_row + 1):
        ws.row_dimensions[row_num].height = row_height

    if include_total:
        total_row = find_total_row(ws, start_row)
        if total_row is not None:
            ws.row_dimensions[total_row].height = row_height

def write_results_to_template(results, template_path, save_path, header_row, sheet_name=None, unit_multiplier=1):
    shutil.copy(template_path, save_path)

    wb = load_workbook(save_path)

    if sheet_name:
        matched_sheet = None

        for real_name in wb.sheetnames:
            if real_name == sheet_name:
                matched_sheet = real_name
                break

        if matched_sheet is None:
            raise ValueError(
                f"找不到指定的 Sheet：{sheet_name}\n实际存在的 Sheet：{wb.sheetnames}"
            )

        ws = wb[matched_sheet]
    else:
        ws = wb.worksheets[0]

    start_row = int(header_row) + 2

    ensure_template_data_rows(
        ws,
        start_row=start_row,
        data_count=len(results),
        min_col=1,
        max_col=10
    )

    apply_production_row_height(
        ws,
        start_row=start_row,
        data_count=len(results)
    )

    for i, row_data in enumerate(results, start=start_row):
        ws[f"B{i}"] = row_data["编号"]

        if row_data["长"] != "":
            ws[f"C{i}"] = cad_round_0(row_data["长"] * unit_multiplier)
        else:
            ws[f"C{i}"] = ""

        if row_data["宽"] != "":
            ws[f"D{i}"] = cad_round_0(row_data["宽"] * unit_multiplier)
        else:
            ws[f"D{i}"] = ""

    wb.save(save_path)
    wb.close()
            
def ask_sheet_mode(parent_window=None):

    win = tk.Toplevel(parent_window)
    win.title("选择输出模式")

    result = {"mode": None}

    tk.Label(
        win,
        text="请选择生产单输出模式：",
        font=("Microsoft JhengHei UI", 10)
    ).pack(pady=(20, 15))

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=10)

    def choose_single():
        result["mode"] = "single"
        win.destroy()

    def choose_multi():
        result["mode"] = "multi"
        win.destroy()

    tk.Button(btn_frame, text="单 Sheet", width=12, command=choose_single).pack(side="left", padx=10)
    tk.Button(btn_frame, text="多 Sheet", width=12, command=choose_multi).pack(side="left", padx=10)

    prepare_dialog(win, parent_window, 320, 150)

    win.wait_window()
    return result["mode"]

def ask_export_unit_multiplier(parent_window=None):
    win = tk.Toplevel(parent_window)
    win.title("选择尺寸单位")

    result = {"multiplier": None}

    tk.Label(
        win,
        text="请选择本次导出尺寸处理方式：",
        font=("Microsoft JhengHei UI", 10, "bold")
    ).pack(pady=(20, 12))

    tk.Label(
        win,
        text="如果 CAD 图面显示 mm，选“不乘”。\n如果 CAD 图面显示 cm，选“乘10”。",
        justify="left"
    ).pack(padx=20, pady=(0, 12))

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=12)

    def choose_1():
        result["multiplier"] = 1
        win.destroy()

    def choose_10():
        result["multiplier"] = 10
        win.destroy()

    def cancel():
        win.destroy()

    tk.Button(btn_frame, text="不乘，直接导出", width=16, command=choose_1).pack(side="left", padx=8)
    tk.Button(btn_frame, text="乘10导出", width=12, command=choose_10).pack(side="left", padx=8)
    tk.Button(btn_frame, text="取消", width=10, command=cancel).pack(side="left", padx=8)

    prepare_dialog(win, parent_window, 460, 210)
    win.wait_window()

    return result["multiplier"]

def ask_target_sheet(sheet_names, parent_window=None):

    win = tk.Toplevel(parent_window)
    win.title("选择目标 Sheet")

    result = {"sheet_name": None}

    tk.Label(
        win,
        text="请选择数据要输出到哪个 Sheet：",
        font=("Microsoft JhengHei UI", 10)
    ).pack(pady=(20, 10))

    combo = ttk.Combobox(win, values=sheet_names, state="readonly", width=28)
    combo.pack(pady=5)

    if sheet_names:
        combo.current(0)

    def confirm():
        value = combo.get()
        if not value:
            messagebox.showerror("错误", "请选择一个 Sheet", parent=win)
            return
        result["sheet_name"] = value
        win.destroy()

    def cancel():
        win.destroy()

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=18)

    tk.Button(btn_frame, text="确定", width=10, command=confirm).pack(side="left", padx=10)
    tk.Button(btn_frame, text="取消", width=10, command=cancel).pack(side="left", padx=10)

    prepare_dialog(win, parent_window, 360, 170)

    win.wait_window()
    return result["sheet_name"]

def ask_small_face_entry(candidate_info, parent_window=None):

    win = tk.Toplevel(parent_window)
    win.title("小面加工确认")

    result = {"choice": None}

    name_text = candidate_info.get("display_name", "")
    type_text = candidate_info.get("object_type", "")

    tk.Label(
        win,
        text="当前选择不是下一组编号。",
        font=("Microsoft JhengHei UI", 10, "bold")
    ).pack(pady=(18, 10))

    tk.Label(
        win,
        text=f"对象名称：{name_text}",
        anchor="w",
        justify="left"
    ).pack(fill="x", padx=20, pady=4)

    tk.Label(
        win,
        text=f"对象类型：{type_text}",
        anchor="w",
        justify="left"
    ).pack(fill="x", padx=20, pady=4)

    tk.Label(
        win,
        text="是否将它作为当前石材的小面加工入口？",
        anchor="w",
        justify="left"
    ).pack(fill="x", padx=20, pady=(10, 8))

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=12)

    def choose_yes():
        result["choice"] = "yes"
        win.destroy()

    def choose_retry():
        result["choice"] = "retry"
        win.destroy()

    def choose_end():
        result["choice"] = "end"
        win.destroy()

    tk.Button(btn_frame, text="是", width=10, command=choose_yes).pack(side="left", padx=8)
    tk.Button(btn_frame, text="否，重选", width=10, command=choose_retry).pack(side="left", padx=8)
    tk.Button(btn_frame, text="结束当前组", width=12, command=choose_end).pack(side="left", padx=8)

    prepare_dialog(win, parent_window, 430, 220)
    win.wait_window()
    return result["choice"]

def assign_small_faces_to_edges(candidates, parent_window=None):
    empty_result = {
        "长边1加工": "",
        "长边2加工": "",
        "短边1加工": "",
        "短边2加工": "",
         "左上角加工":"", 
         "右上角加工":"", 
         "左下角加工": "", 
         "右下角加工": ""
    }

    if not candidates:
        return {"action": "confirm", "data": empty_result}

    win = tk.Toplevel(parent_window)
    win.title("小面加工归类")

    tk.Label(
        win,
        text="请为本组小面加工选择归属边：",
        font=("Microsoft JhengHei UI", 10, "bold")
    ).pack(pady=(15, 10))

    table_frame = tk.Frame(win)
    table_frame.pack(fill="both", expand=True, padx=12, pady=5)

    headers = ["序号", "对象名称", "对象类型", "归属边"]
    for col, text in enumerate(headers):
        tk.Label(
            table_frame,
            text=text,
            font=("Microsoft JhengHei UI", 9, "bold"),
            relief="groove",
            width=18 if col != 1 else 24
        ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

    options = ["不录入", "长边1加工", "长边2加工", "短边1加工", "短边2加工","左上角加工", "右上角加工", "左下角加工", "右下角加工"]
    choice_vars = []

    for i, item in enumerate(candidates, start=1):
        tk.Label(table_frame, text=str(i), relief="groove", width=18).grid(
            row=i, column=0, sticky="nsew", padx=1, pady=1
        )
        tk.Label(table_frame, text=item.get("display_name", ""), relief="groove", width=24).grid(
            row=i, column=1, sticky="nsew", padx=1, pady=1
        )
        tk.Label(table_frame, text=item.get("object_type", ""), relief="groove", width=18).grid(
            row=i, column=2, sticky="nsew", padx=1, pady=1
        )

        var = tk.StringVar(value="不录入")
        combo = ttk.Combobox(
            table_frame,
            textvariable=var,
            values=options,
            state="readonly",
            width=16
        )
        combo.grid(row=i, column=3, sticky="nsew", padx=1, pady=1)

        choice_vars.append((item, var))

    final_result = {"action": "cancel", "data": None}

    def build_assignments():
        temp = {
            "长边1加工": "",
            "长边2加工": "",
            "短边1加工": "",
            "短边2加工": "",
            "左上角加工": "",
            "右上角加工": "",
            "左下角加工": "",
            "右下角加工": ""
        }

        for item, var in choice_vars:
            slot = var.get()
            if slot == "不录入":
                continue

            value = item.get("display_name", "")

            if temp[slot]:
                temp[slot] += "、" + value
            else:
                temp[slot] = value

        return temp

    def confirm():
        final_result["action"] = "confirm"
        final_result["data"] = build_assignments()
        win.destroy()

    def go_back():
        final_result["action"] = "back"
        final_result["data"] = None
        win.destroy()

    def cancel():
        final_result["action"] = "cancel"
        final_result["data"] = None
        win.destroy()

    btn_frame = tk.Frame(win)
    btn_frame.pack(pady=12)

    tk.Button(btn_frame, text="确定", width=10, command=confirm).pack(side="left", padx=8)
    tk.Button(btn_frame, text="返回继续选", width=12, command=go_back).pack(side="left", padx=8)
    tk.Button(btn_frame, text="取消", width=10, command=cancel).pack(side="left", padx=8)

    prepare_dialog(win, parent_window, 820, 320)

    win.wait_window()
    return final_result

def apply_small_face_assignments_to_row(row, assignments):
    if not assignments:
        return

    for key in ["长边1加工", "长边2加工", "短边1加工", "短边2加工",
                "左上角加工", "右上角加工", "左下角加工", "右下角加工"]:
        row[key] = assignments.get(key, "")
        
def collect_small_face_types(results):
    fields = ["长边1加工", "长边2加工", "短边1加工", "短边2加工",
              "左上角加工", "右上角加工", "左下角加工", "右下角加工"]
    type_set = set()

    for row in results:
        for field in fields:
            raw_value = str(row.get(field, "")).strip()
            if not raw_value:
                continue

            parts = [p.strip() for p in raw_value.split("、") if p.strip()]
            for part in parts:
                type_set.add(part)

    return sorted(type_set)

def ask_small_face_symbol_mapping(process_names, symbol_options, parent_window=None):
    win = tk.Toplevel(parent_window)
    win.title("选择小面加工对应图示")

    tk.Label(
        win,
        text="请为本次出现过的小面加工类型选择对应图示：",
        font=("Microsoft JhengHei UI", 10, "bold")
    ).pack(pady=(15, 10))

    table_frame = tk.Frame(win)
    table_frame.pack(fill="both", expand=True, padx=12, pady=5)

    headers = ["序号", "小面加工类型", "对应图示"]
    widths = [8, 28, 20]

    for col, text in enumerate(headers):
        tk.Label(
            table_frame,
            text=text,
            font=("Microsoft JhengHei UI", 9, "bold"),
            relief="groove",
            width=widths[col]
        ).grid(row=0, column=col, sticky="nsew", padx=1, pady=1)

    display_options = ["不绘制"] + [symbol_name for symbol_id, symbol_name in symbol_options]
    name_to_id = {symbol_name: symbol_id for symbol_id, symbol_name in symbol_options}

    choice_vars = []

    for i, process_name in enumerate(process_names, start=1):
        tk.Label(table_frame, text=str(i), relief="groove", width=widths[0]).grid(
            row=i, column=0, sticky="nsew", padx=1, pady=1
        )

        tk.Label(table_frame, text=process_name, relief="groove", width=widths[1], anchor="w").grid(
            row=i, column=1, sticky="nsew", padx=1, pady=1
        )

        var = tk.StringVar(value=display_options[0])
        combo = ttk.Combobox(
            table_frame,
            textvariable=var,
            values=display_options,
            state="readonly",
            width=widths[2] - 2
        )
        combo.grid(row=i, column=2, sticky="nsew", padx=1, pady=1)

        choice_vars.append((process_name, var))

    result = {"mapping": None}

    def confirm():
        mapping = {}
        for process_name, var in choice_vars:
            selected_name = var.get().strip()

            if selected_name == "不绘制":
                mapping[process_name] = "不绘制"
            else:
                mapping[process_name] = name_to_id.get(selected_name, "不绘制")

        result["mapping"] = mapping
        win.destroy()

    def cancel():
        win.destroy()

    win.protocol("WM_DELETE_WINDOW", cancel)

    btn_frame = tk.Frame(win)
    btn_frame.pack(fill="x", pady=(8, 15))

    tk.Button(btn_frame, text="确定", width=10, command=confirm).pack(side="left", padx=20)
    tk.Button(btn_frame, text="取消", width=10, command=cancel).pack(side="right", padx=20)

    prepare_dialog(win, parent_window, 620, 360)

    win.wait_window()
    return result["mapping"]
    
def export_results_to_production_sheet(results, process_symbol_map, config, parent_window=None):
    if not results:
        messagebox.showwarning("提示", "没有数据可导出")
        return

    template_path = filedialog.askopenfilename(
        title="请选择生产单模板",
        filetypes=[("Excel文件", "*.xlsx")]
    )
    if not template_path:
        return

    save_path = filedialog.asksaveasfilename(
        title="请选择保存位置",
        defaultextension=".xlsx",
        filetypes=[("Excel文件", "*.xlsx")]
    )
    if not save_path:
        return

    sheet_mode = ask_sheet_mode(parent_window)
    if not sheet_mode:
        return

    selected_sheet = None

    if sheet_mode == "multi":
        try:
            wb_preview = load_workbook(template_path, read_only=True)
            sheet_names = wb_preview.sheetnames
            wb_preview.close()
        except Exception as e:
            messagebox.showerror("错误", f"读取模板 Sheet 失败：\n{e}")
            return

        if not sheet_names:
            messagebox.showerror("错误", "该模板没有可用的 Sheet")
            return

        selected_sheet = ask_target_sheet(sheet_names, parent_window)
        if not selected_sheet:
            return

    header_row = simpledialog.askstring(
        "输入行号",
        "请输入表头上一行的行号（例如输入35 → 数据从37开始）：",
        parent=parent_window
    )

    if not header_row:
        return

    if not header_row.isdigit():
        messagebox.showerror("错误", "请输入数字行号")
        return
    
    unit_multiplier = ask_export_unit_multiplier(parent_window)

    if unit_multiplier is None:
        return
    
    shape_col_letter = SMALL_FACE_COL
    
    try:
        write_results_to_template(
            results,
            template_path,
            save_path,
            int(header_row),
            sheet_name=selected_sheet,
            unit_multiplier=unit_multiplier
        )

        start_row = int(header_row) + 2

        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = True

        wb = excel.Workbooks.Open(save_path)
        
        excel.CalculateFull()
        time.sleep(0.2)

        if selected_sheet:
            ws = wb.Worksheets(selected_sheet)
        else:
            ws = wb.Worksheets(1)

        draw_small_face_blocks_to_sheet(
            ws,
            excel,
            results,
            start_row,
            shape_col_letter,
            process_symbol_map,
            config
        )

        wb.Save()

        archive_path = archive_progress_after_export(
            save_path=save_path,
            sheet_name=selected_sheet
        )

        backup_msg = ""
        if archive_path:
            backup_msg = f"\n\n录入记录已备份到：\n{archive_path}"

        if selected_sheet:
            messagebox.showinfo(
                "完成",
                f"已成功导出到：\n{save_path}\n\n目标 Sheet：{selected_sheet}\n图示输出列：{shape_col_letter}{backup_msg}"
            )
        else:
            messagebox.showinfo(
                "完成",
                f"已成功导出到：\n{save_path}\n\n图示输出列：{shape_col_letter}{backup_msg}"
            )
        
        clear_progress()

    except Exception as e:
        messagebox.showerror(
            "错误",
            f"导出失败，所选模板可能不是 openpyxl 可读取的标准 Excel 文件。\n\n详细信息：\n{e}"
        )
        
def handle_post_width_stage(doc, panel):
    panel.set_step("请选择下一对象（编号 / 小面加工）")

    while True:
        print("下一步：请选择下一组编号，或当前组的小面加工对象")
        obj = get_entity_by_click(doc, "请选择下一组编号，或当前组的小面加工对象", panel.root)

        if obj is None:
            return {"action": "end_group"}

        kind, payload = classify_post_width_entity(obj)

        if kind == "invalid":
            print(payload)
            panel.set_status(payload)
            continue

        if kind == "next_code":
            panel.clear_status()
            return {
                "action": "next_code",
                "code": payload
            }

        if kind == "non_code":
            panel.set_status(f"已选择非编号对象：{payload['display_name']} ({payload['object_type']})")
            choice = ask_small_face_entry(payload, panel.root)

            if choice == "yes":
                panel.clear_status()
                return {
                    "action": "small_face",
                    "candidate": payload
                }

            if choice == "retry":
                panel.set_status("请重新选择下一对象。")
                continue

            if choice == "end":
                panel.clear_status()
                return {"action": "end_group"}

def process_post_width(doc, panel, small_face_candidates):
    panel.set_step("请选择下一对象（编号 / 小面加工）")
    entered_small_face_mode = False

    while True:
        print("下一步：请选择下一组编号，或当前组的小面加工对象")
        time.sleep(0.2)

        try:
            obj = get_entity_by_click(
                doc,
                "请选择下一组编号，或当前组的小面加工对象",
                panel.root,
                allow_finish_by_empty=entered_small_face_mode
            )
            if obj is None and entered_small_face_mode:
                panel.set_status("已结束小面加工选择")
                return None

        except DiscardCurrentGroup:
            panel.set_status("已放弃当前组，请重新录入。")
            panel.clear_temp()
            raise

        except EndWholeInput:
            panel.clear_status()
            return "__END_AFTER_CURRENT__"

        except Exception:
            if entered_small_face_mode:
                panel.set_status("已结束小面加工选择")
                return None
            else:
                raise

        kind, payload = classify_post_width_entity(obj)

        if kind == "invalid":
            print(payload)
            panel.set_status(payload)
            continue

        if kind == "next_code":
            msg = f"检测到下一组编号：{payload}"
            print(msg)
            panel.set_status(msg)
            return payload

        if kind == "non_code":
            if entered_small_face_mode:
                if is_duplicate_small_face(payload, small_face_candidates):
                    msg = f"重复选择了同一个小面加工对象：{payload['display_name']}"
                    print(msg)
                    panel.set_status(msg)
                    continue

                small_face_candidates.append(payload)
                msg = f"已加入小面加工候选：{payload['display_name']}（当前共 {len(small_face_candidates)} 个）"
                print(msg)
                panel.set_status(msg)
                continue

            panel.set_status(
                f"已选择非编号对象：{payload['display_name']} ({payload['object_type']})"
            )
            choice = ask_small_face_entry(payload, panel.root)

            if choice == "yes":
                if is_duplicate_small_face(payload, small_face_candidates):
                    msg = f"重复选择了同一个小面加工对象：{payload['display_name']}"
                    print(msg)
                    panel.set_status(msg)
                    continue

                small_face_candidates.append(payload)
                entered_small_face_mode = True
                msg = f"已加入小面加工候选：{payload['display_name']}"
                print(msg)
                panel.set_status(msg)
                continue

            elif choice == "retry":
                panel.set_status("当前对象未加入，请重新选择。")
                continue

            elif choice == "end":
                panel.clear_status()
                return None
            
COLOR_MAP = {
    "black": 0x000000,
    "red": 0x0000FF,
    "green": 0x00FF00,
    "blue": 0xFF0000,
    "white": 0xFFFFFF,
    "yellow": 0x00FFFF
}


def load_config(config_path):
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)
    
def normalize_shape_cfg(shape_cfg):
    if not isinstance(shape_cfg, dict):
        return shape_cfg

    category = shape_cfg.get("category", "basic")
    if category != "basic":
        return shape_cfg

    if "type" in shape_cfg:
        return shape_cfg

    if "shape" in shape_cfg:
        merged = {
            "category": "basic"
        }

        shape_part = shape_cfg.get("shape", {})
        style_part = shape_cfg.get("style", {})

        merged.update(shape_part)

        if "line" in style_part:
            merged["line"] = style_part["line"]
        if "fill" in style_part:
            merged["fill"] = style_part["fill"]

        for k, v in shape_cfg.items():
            if k not in ["shape", "style"]:
                merged[k] = v

        return merged

    return shape_cfg


def parse_color_value(color_value):
    if isinstance(color_value, str):
        return COLOR_MAP.get(color_value.lower(), 0x000000)

    if isinstance(color_value, (list, tuple)) and len(color_value) == 3:
        r, g, b = color_value
        try:
            r, g, b = int(r), int(g), int(b)
            # Excel COM 的 RGB 是 BGR 顺序整数
            return (b << 16) | (g << 8) | r
        except Exception:
            return 0x000000

    return 0x000000


def apply_shape_style(shape, shape_cfg):
    shape_cfg = normalize_shape_cfg(shape_cfg)

    line_cfg = shape_cfg.get("line", {})
    fill_cfg = shape_cfg.get("fill", {})

    line_visible = line_cfg.get("visible", True)
    shape.Line.Visible = -1 if line_visible else 0
    shape.Line.Weight = line_cfg.get("weight", 1)

    line_color_value = line_cfg.get("color", "black")
    shape.Line.ForeColor.RGB = parse_color_value(line_color_value)

    fill_visible = fill_cfg.get("visible", False)
    shape.Fill.Visible = -1 if fill_visible else 0

    fill_color_value = fill_cfg.get("color", "white")
    shape.Fill.ForeColor.RGB = parse_color_value(fill_color_value)


def get_shape_size_pt(excel, shape_cfg, symbol_index=None):
    category = shape_cfg.get("category", "basic")

    # ===== basic =====
    if category == "basic":
        shape_cfg = normalize_shape_cfg(shape_cfg)
        shape_type = shape_cfg["type"]

        if shape_type == "rectangle":
            if "width_cm" in shape_cfg and "height_cm" in shape_cfg:
                w = excel.CentimetersToPoints(shape_cfg["width_cm"])
                h = excel.CentimetersToPoints(shape_cfg["height_cm"])
            elif "size_cm" in shape_cfg:
                s = excel.CentimetersToPoints(shape_cfg["size_cm"])
                w, h = s, s
            else:
                raise ValueError("rectangle 缺少 width_cm/height_cm 或 size_cm")

        elif shape_type == "oval":
            d = excel.CentimetersToPoints(shape_cfg["diameter_cm"])
            w, h = d, d
        
        elif shape_type == "chevron":
            if "width_cm" in shape_cfg and "height_cm" in shape_cfg:
                w = excel.CentimetersToPoints(shape_cfg["width_cm"])
                h = excel.CentimetersToPoints(shape_cfg["height_cm"])
            elif "size_cm" in shape_cfg:
                s = excel.CentimetersToPoints(shape_cfg["size_cm"])
                w, h = s, s
            else:
                raise ValueError("chevron 缺少 width_cm/height_cm 或 size_cm")

        elif shape_type == "triangle_equilateral":
            side = excel.CentimetersToPoints(shape_cfg["side_cm"])
            w = side
            h = side * math.sqrt(3) / 2
        
        elif shape_type == "circle_cross":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s

        elif shape_type == "cross":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s

        elif shape_type == "plus":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s

        elif shape_type == "diamond":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s
            
        elif shape_type == "5star":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s
            
        elif shape_type == "6star":
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s
        
        elif shape_type in ["triangle_right_isosceles", "right_triangle"]:
            leg = excel.CentimetersToPoints(shape_cfg["leg_cm"])
            w, h = leg, leg

        else:
            raise ValueError(f"未知图形类型: {shape_type}")

        return w, h

    # text
    elif category == "text":
        if "width_cm" in shape_cfg and "height_cm" in shape_cfg:
            w = excel.CentimetersToPoints(shape_cfg["width_cm"])
            h = excel.CentimetersToPoints(shape_cfg["height_cm"])
        elif "size_cm" in shape_cfg:
            s = excel.CentimetersToPoints(shape_cfg["size_cm"])
            w, h = s, s
        else:
            raise ValueError("text 缺少 width_cm/height_cm 或 size_cm")

        return w, h
    
    # omposite
    elif category == "composite":
        meta = shape_cfg.get("meta", {})
        size_hint = meta.get("size_hint", {})

        if "width_cm" in size_hint and "height_cm" in size_hint:
            w = excel.CentimetersToPoints(size_hint["width_cm"])
            h = excel.CentimetersToPoints(size_hint["height_cm"])
            return w, h

        if "width" in size_hint and "height" in size_hint:
            w = excel.CentimetersToPoints(size_hint["width"])
            h = excel.CentimetersToPoints(size_hint["height"])
            return w, h

        if symbol_index is None:
            raise ValueError("composite 自动估算尺寸时需要 symbol_index")

        children = shape_cfg.get("children", [])
        if not children:
            raise ValueError("composite 缺少 children")

        min_left = float("inf")
        max_right = float("-inf")
        min_top = float("inf")
        max_bottom = float("-inf")

        for child in children:
            if "shape" in child:
                child_cfg = child["shape"]
            elif "ref" in child:
                ref_id = child["ref"]
                child_cfg = symbol_index.get(ref_id)
                if not child_cfg:
                    raise ValueError(f"composite 引用了不存在的 symbol: {ref_id}")
            else:
                raise ValueError("child 必须包含 shape 或 ref")

            child_w, child_h = get_shape_size_pt(excel, child_cfg, symbol_index)

            dx = child.get("offset", {}).get("x", 0)
            dy = child.get("offset", {}).get("y", 0)

            dx_pt = excel.CentimetersToPoints(dx)
            dy_pt = excel.CentimetersToPoints(dy)

            left = dx_pt - child_w / 2
            right = dx_pt + child_w / 2
            top = dy_pt - child_h / 2
            bottom = dy_pt + child_h / 2

            min_left = min(min_left, left)
            max_right = max(max_right, right)
            min_top = min(min_top, top)
            max_bottom = max(max_bottom, bottom)

        w = max_right - min_left
        h = max_bottom - min_top
        return w, h

    else:
        raise ValueError(f"未知 category: {category}")

def apply_textbox_style(textbox, text_cfg):
    line_cfg = text_cfg.get("line", {})
    fill_cfg = text_cfg.get("fill", {})

    line_visible = line_cfg.get("visible", False)
    textbox.Line.Visible = -1 if line_visible else 0
    if line_visible:
        textbox.Line.Weight = line_cfg.get("weight", 0.25)
        line_color_value = line_cfg.get("color", "black")
        textbox.Line.ForeColor.RGB = parse_color_value(line_color_value)

    fill_visible = fill_cfg.get("visible", False)
    textbox.Fill.Visible = -1 if fill_visible else 0
    if fill_visible:
        fill_color_value = fill_cfg.get("color", "white")
        textbox.Fill.ForeColor.RGB = parse_color_value(fill_color_value)

    content = str(text_cfg.get("content", ""))
    font_size = text_cfg.get("font_size", 7)
    bold = bool(text_cfg.get("bold", False))

    try:
        textbox.TextFrame2.AutoSize = 0
    except Exception:
        pass

    try:
        textbox.TextFrame.AutoSize = False
    except Exception:
        pass

    try:
        textbox.TextFrame2.WordWrap = False
    except Exception:
        pass

    textbox.TextFrame2.TextRange.Text = content
    textbox.TextFrame2.TextRange.Font.Size = font_size
    textbox.TextFrame2.TextRange.Font.Bold = -1 if bold else 0

    text_color_value = text_cfg.get("font_color", line_cfg.get("color", "black"))
    textbox.TextFrame2.TextRange.Font.Fill.ForeColor.RGB = parse_color_value(text_color_value)

    textbox.TextFrame2.VerticalAnchor = 3
    textbox.TextFrame2.TextRange.ParagraphFormat.Alignment = 2

    try:
        textbox.TextFrame.MarginLeft = 0
        textbox.TextFrame.MarginRight = 0
        textbox.TextFrame.MarginTop = 0
        textbox.TextFrame.MarginBottom = 0
    except Exception:
        pass

    try:
        textbox.TextFrame2.MarginLeft = 0
        textbox.TextFrame2.MarginRight = 0
        textbox.TextFrame2.MarginTop = 0
        textbox.TextFrame2.MarginBottom = 0
    except Exception:
        pass

def create_shape_from_config(ws, excel, shape_cfg, center_x, center_y, name=None, symbol_index=None):
    category = shape_cfg.get("category", "basic")

    # basic
    if category == "basic":
        shape_cfg = normalize_shape_cfg(shape_cfg)
        shape_type = shape_cfg["type"]
        w, h = get_shape_size_pt(excel, shape_cfg, symbol_index)

        left = center_x - w / 2
        top = center_y - h / 2

        if shape_type == "rectangle":
            shape = ws.Shapes.AddShape(1, left, top, w, h)

        elif shape_type == "oval":
            shape = ws.Shapes.AddShape(9, left, top, w, h)
            
        elif shape_type == "chevron":
            shape = ws.Shapes.AddShape(52, left, top, w, h)

        elif shape_type == "triangle_equilateral":
            shape = ws.Shapes.AddShape(7, left, top, w, h)

        elif shape_type in ["triangle_right_isosceles", "right_triangle"]:
            shape = ws.Shapes.AddShape(8, left, top, w, h)  # 8 = 直角三角形
        
        elif shape_type == "diamond":
            shape = ws.Shapes.AddShape(63, left, top, w, h)

        elif shape_type == "cross":
            shape = ws.Shapes.AddShape(165, left, top, w, h)

        elif shape_type == "plus":
            shape = ws.Shapes.AddShape(163, left, top, w, h)
            
        elif shape_type == "5star":
            shape = ws.Shapes.AddShape(92, left, top, w, h)

        elif shape_type == "6star":
            shape = ws.Shapes.AddShape(147, left, top, w, h)

        elif shape_type == "circle_cross":
            shape = ws.Shapes.AddShape(77, left, top, w, h)

        else:
            raise ValueError(f"未知图形类型: {shape_type}")

        apply_shape_style(shape, shape_cfg)
        
        adjustments = shape_cfg.get("adjustments", [])
        for i, val in enumerate(adjustments, start=1):
            try:
                shape.Adjustments[i] = float(val)
            except Exception:
                pass

        rotation = shape_cfg.get("rotation", 0)
        try:
            shape.Rotation = float(rotation)
        except Exception:
            pass

        if name:
            try:
                shape.Name = name
            except Exception:
                pass

        return shape


    # text
    elif category == "text":
        w, h = get_shape_size_pt(excel, shape_cfg, symbol_index)

        left = center_x - w / 2
        top = center_y - h / 2

        textbox = ws.Shapes.AddTextbox(1, left, top, w, h)

        apply_textbox_style(textbox, shape_cfg)

        textbox.Left = left
        textbox.Top = top
        textbox.Width = w
        textbox.Height = h

        if name:
            try:
                textbox.Name = name
            except Exception:
                pass

        return textbox
    
    # composite
    elif category == "composite":
        if symbol_index is None:
            raise ValueError("composite 绘制需要 symbol_index")

        children = shape_cfg.get("children", [])
        if not children:
            raise ValueError("composite 缺少 children")

        created_names = []

        for idx, child in enumerate(children, start=1):
            if "shape" in child:
                child_cfg = child["shape"]
                child_id_for_name = child_cfg.get("type", f"child{idx}")
            elif "ref" in child:
                ref_id = child["ref"]
                child_cfg = symbol_index.get(ref_id)

                if not child_cfg:
                    raise ValueError(f"composite 引用了不存在的 symbol: {ref_id}")

                child_id_for_name = ref_id
            else:
                raise ValueError("child 必须包含 shape 或 ref")

            child_offset = child.get("offset", {})
            dx = child_offset.get("x", 0)
            dy = child_offset.get("y", 0)

            child_cx = center_x + excel.CentimetersToPoints(dx)
            child_cy = center_y + excel.CentimetersToPoints(dy)

            child_name = None
            if name:
                child_name = f"{name}_child{idx}_{child_id_for_name}"

            child_shape = create_shape_from_config(
                ws=ws,
                excel=excel,
                shape_cfg=child_cfg,
                center_x=child_cx,
                center_y=child_cy,
                name=child_name,
                symbol_index=symbol_index
            )
            created_names.append(child_shape.Name)

        grouped = ws.Shapes.Range(created_names).Group()

        if name:
            try:
                grouped.Name = name
            except Exception:
                pass

        return grouped

    else:
        raise ValueError(f"未知 category: {category}")


def get_cell_center(cell):
    cx = cell.Left + cell.Width / 2
    cy = cell.Top + cell.Height / 2
    return cx, cy


def get_symbol_center_by_side(base_cx, base_cy, base_w, base_h, symbol_w, symbol_h, side, gap_pt):
    if side == "top":
        cx = base_cx
        cy = base_cy - base_h / 2 - symbol_h / 2 - gap_pt
    elif side == "bottom":
        cx = base_cx
        cy = base_cy + base_h / 2 + symbol_h / 2 + gap_pt
    elif side == "left":
        cx = base_cx - base_w / 2 - symbol_w / 2 - gap_pt
        cy = base_cy
    elif side == "right":
        cx = base_cx + base_w / 2 + symbol_w / 2 + gap_pt
        cy = base_cy
    else:
        raise ValueError(f"未知边: {side}")

    return cx, cy

def get_symbol_center_by_corner(base_cx, base_cy, base_w, base_h, symbol_w, symbol_h, corner, gap_pt):
    if corner == "top_left":
        cx = base_cx - base_w / 2 - symbol_w / 2 - gap_pt
        cy = base_cy - base_h / 2 - symbol_h / 2 - gap_pt
    elif corner == "top_right":
        cx = base_cx + base_w / 2 + symbol_w / 2 + gap_pt
        cy = base_cy - base_h / 2 - symbol_h / 2 - gap_pt
    elif corner == "bottom_left":
        cx = base_cx - base_w / 2 - symbol_w / 2 - gap_pt
        cy = base_cy + base_h / 2 + symbol_h / 2 + gap_pt
    elif corner == "bottom_right":
        cx = base_cx + base_w / 2 + symbol_w / 2 + gap_pt
        cy = base_cy + base_h / 2 + symbol_h / 2 + gap_pt
    else:
        raise ValueError(f"未知角点: {corner}")

    return cx, cy

def get_multi_symbol_offset(side, symbol_w, symbol_h, index, total_count, extra_gap_pt):
    if total_count <= 1:
        return 0, 0

    if total_count == 2:
        if side in ["top", "bottom"]:
            shift = symbol_w / 2 + extra_gap_pt / 2
            return (-shift, 0) if index == 0 else (shift, 0)

        if side in ["left", "right"]:
            shift = symbol_h / 2 + extra_gap_pt / 2
            return (0, -shift) if index == 0 else (0, shift)

    if side in ["top", "bottom"]:
        spacing = symbol_w + extra_gap_pt
        start = -spacing * (total_count - 1) / 2
        dx = start + index * spacing
        return dx, 0

    if side in ["left", "right"]:
        spacing = symbol_h + extra_gap_pt
        start = -spacing * (total_count - 1) / 2
        dy = start + index * spacing
        return 0, dy

    return 0, 0


def get_symbol_cfg_by_id(config, symbol_id):
    symbol_index = build_symbol_index(config)
    if symbol_id in symbol_index:
        return symbol_index[symbol_id]
    raise ValueError(f"找不到 symbol id: {symbol_id}")

def build_symbol_index(config):
    symbol_index = {}
    for item in config.get("symbols", []):
        symbol_id = str(item.get("id", "")).strip()
        if symbol_id:
            symbol_index[symbol_id] = item
    return symbol_index

def get_symbol_options_from_config(config):
    options = []

    for item in config.get("symbols", []):
        symbol_id = str(item.get("id", "")).strip()
        symbol_name = str(item.get("name", "")).strip()

        if not symbol_id:
            continue

        if not symbol_name:
            symbol_name = symbol_id

        options.append((symbol_id, symbol_name))

    return options

def group_shapes_by_names(ws, shape_names, group_name=None):
    shape_range = ws.Shapes.Range(shape_names)
    grouped = shape_range.Group()

    if group_name:
        try:
            grouped.Name = group_name
        except Exception:
            pass

    return grouped

def delete_shapes_by_prefix(ws, prefixes):
    to_delete = []
    for i in range(1, ws.Shapes.Count + 1):
        try:
            shp = ws.Shapes.Item(i)
            shp_name = shp.Name
            if any(shp_name.startswith(prefix) for prefix in prefixes):
                to_delete.append(shp_name)
        except Exception:
            pass

    for name in to_delete:
        try:
            ws.Shapes(name).Delete()
        except Exception:
            pass
        
def build_edge_symbol_map(row, process_symbol_map):
    field_to_side = {
        "长边1加工": "top",
        "长边2加工": "bottom",
        "短边1加工": "left",
        "短边2加工": "right",
        "左上角加工": "top_left",
        "右上角加工": "top_right",
        "左下角加工": "bottom_left",
        "右下角加工": "bottom_right"
    }

    edge_symbol_map = {
        "top": [],
        "bottom": [],
        "left": [],
        "right": [],
        "top_left": [],
        "top_right": [],
        "bottom_left": [],
        "bottom_right": []
    }

    for field, side in field_to_side.items():
        raw_value = str(row.get(field, "")).strip()
        if not raw_value:
            continue

        parts = [p.strip() for p in raw_value.split("、") if p.strip()]
        if not parts:
            continue

        for process_name in parts:
            symbol_id = process_symbol_map.get(process_name)
            if symbol_id:
                edge_symbol_map[side].append(symbol_id)

    return edge_symbol_map

def draw_small_face_block_in_cell(ws, excel, cell, edge_symbol_map, config, gap_cm=0.03, multi_gap_cm=0.03):
    cell_addr = cell.Address.replace("$", "")
    symbol_index = build_symbol_index(config)

    prefix = f"{ws.Name}_"

    delete_shapes_by_prefix(ws, [
        f"{prefix}BaseRect_{cell_addr}",
        f"{prefix}Top_{cell_addr}_",
        f"{prefix}Bottom_{cell_addr}_",
        f"{prefix}Left_{cell_addr}_",
        f"{prefix}Right_{cell_addr}_",
        f"{prefix}top_left_{cell_addr}_",
        f"{prefix}top_right_{cell_addr}_",
        f"{prefix}bottom_left_{cell_addr}_",
        f"{prefix}bottom_right_{cell_addr}_",
        f"{prefix}SmallFaceBlock_{cell_addr}"
    ])

    base_rect_cfg = config["base_rect"]

    base_cx, base_cy = get_cell_center(cell)
    base_w, base_h = get_shape_size_pt(excel, base_rect_cfg, symbol_index)
    gap_pt = excel.CentimetersToPoints(gap_cm)
    multi_gap_pt = excel.CentimetersToPoints(multi_gap_cm)

    created_names = []

    base_shape = create_shape_from_config(
        ws, excel, base_rect_cfg, base_cx, base_cy,
        name=f"{ws.Name}_BaseRect_{cell_addr}",
        symbol_index=symbol_index
    )
    created_names.append(base_shape.Name)

    for side in ["top", "bottom", "left", "right"]:
        symbol_ids = edge_symbol_map.get(side, [])
        if not symbol_ids:
            continue

        total_count = len(symbol_ids)

        for idx, symbol_id in enumerate(symbol_ids):
            symbol_cfg = symbol_index.get(symbol_id)
            if not symbol_cfg:
                raise ValueError(f"找不到 symbol id: {symbol_id}")

            symbol_w, symbol_h = get_shape_size_pt(excel, symbol_cfg, symbol_index)

            cx, cy = get_symbol_center_by_side(
                base_cx, base_cy,
                base_w, base_h,
                symbol_w, symbol_h,
                side,
                gap_pt
            )

            dx, dy = get_multi_symbol_offset(
                side,
                symbol_w,
                symbol_h,
                idx,
                total_count,
                multi_gap_pt
            )
            cx += dx
            cy += dy

            shape_name = f"{ws.Name}_{side.capitalize()}_{cell_addr}_{idx + 1}_{symbol_id}"

            symbol_shape = create_shape_from_config(
                ws=ws,
                excel=excel,
                shape_cfg=symbol_cfg,
                center_x=cx,
                center_y=cy,
                name=shape_name,
                symbol_index=symbol_index
            )
            created_names.append(symbol_shape.Name)

    for corner in ["top_left", "top_right", "bottom_left", "bottom_right"]:
        symbol_ids = edge_symbol_map.get(corner, [])
        if not symbol_ids:
            continue

        for idx, symbol_id in enumerate(symbol_ids):
            symbol_cfg = symbol_index.get(symbol_id)
            if not symbol_cfg:
                raise ValueError(f"找不到 symbol id: {symbol_id}")

            symbol_w, symbol_h = get_shape_size_pt(excel, symbol_cfg, symbol_index)

            cx, cy = get_symbol_center_by_corner(
                base_cx, base_cy,
                base_w, base_h,
                symbol_w, symbol_h,
                corner,
                gap_pt
            )

            shape_name = f"{ws.Name}_{corner}_{cell_addr}_{idx + 1}_{symbol_id}"

            symbol_shape = create_shape_from_config(
                ws=ws,
                excel=excel,
                shape_cfg=symbol_cfg,
                center_x=cx,
                center_y=cy,
                name=shape_name,
                symbol_index=symbol_index
            )
            created_names.append(symbol_shape.Name)

    if created_names:
        grouped_shape = group_shapes_by_names(
            ws,
            created_names,
            group_name=f"{ws.Name}_SmallFaceBlock_{cell_addr}"
        )

        try:
            grouped_shape.Placement = 1  
        except Exception:
            pass

        return grouped_shape

    return None
            
def row_has_small_face(row):
    fields = ["长边1加工", "长边2加工", "短边1加工", "短边2加工",
              "左上角加工", "右上角加工", "左下角加工", "右下角加工"]

    for field in fields:
        if str(row.get(field, "")).strip():
            return True

    return False

def draw_small_face_blocks_to_sheet(ws, excel, results, start_row, shape_col_letter, process_symbol_map, config):
    for idx, row in enumerate(results):
        target_row = start_row + idx

        if not row_has_small_face(row):
            print(f"[绘图] 第 {idx + 1} 条 -> 第 {target_row} 行：无小面加工，跳过")
            continue

        edge_symbol_map = build_edge_symbol_map(row, process_symbol_map)

        if not any(edge_symbol_map.values()):
            print(f"[绘图] 第 {idx + 1} 条 -> 第 {target_row} 行：无可绘制图示，跳过")
            continue

        cell_addr = f"{shape_col_letter}{target_row}"
        cell = ws.Range(cell_addr)

        print(f"[绘图] 第 {idx + 1} 条 -> {cell_addr} -> {edge_symbol_map}")

        draw_small_face_block_in_cell(
            ws,
            excel,
            cell,
            edge_symbol_map,
            config
        )

def ensure_backup_dir():
    os.makedirs(BACKUP_DIR, exist_ok=True)


def save_progress(results):
    ensure_backup_dir()
    with open(PROGRESS_JSON_PATH, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


def load_progress():
    if not os.path.exists(PROGRESS_JSON_PATH):
        return []

    with open(PROGRESS_JSON_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def archive_progress_after_export(save_path=None, sheet_name=None):
    if not os.path.exists(PROGRESS_JSON_PATH):
        return None

    ensure_backup_dir()
    os.makedirs(PROGRESS_HISTORY_DIR, exist_ok=True)

    timestamp = time.strftime("%Y%m%d_%H%M%S")

    safe_output_name = ""
    if save_path:
        base_name = os.path.splitext(os.path.basename(str(save_path)))[0]
        safe_output_name = re.sub(r'[\\/:*?"<>|\s]+', "_", base_name).strip("_")[:60]

    filename_parts = ["progress", timestamp]
    if safe_output_name:
        filename_parts.append(safe_output_name)

    archive_path = os.path.join(
        PROGRESS_HISTORY_DIR,
        "_".join(filename_parts) + ".json"
    )

    shutil.copy2(PROGRESS_JSON_PATH, archive_path)
    shutil.copy2(PROGRESS_JSON_PATH, LAST_SUCCESS_JSON_PATH)

    meta_path = archive_path.replace(".json", "_meta.json")
    meta_data = {
        "exported_at": time.strftime("%Y-%m-%d %H:%M:%S"),
        "save_path": save_path,
        "sheet_name": sheet_name
    }
    try:
        with open(meta_path, "w", encoding="utf-8") as f:
            json.dump(meta_data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return archive_path


def clear_progress():
    if os.path.exists(PROGRESS_JSON_PATH):
        os.remove(PROGRESS_JSON_PATH)

def main():
    print("正在连接 AutoCAD...")
    acad, doc = get_autocad_document()
    print("已连接到 AutoCAD。")

    panel = FloatingPanel()
    results = []
    pending_code = None

    if os.path.exists(PROGRESS_JSON_PATH):
        answer = messagebox.askyesno(
            "恢复进度",
            "检测到上次未完成的录入进度。\n\n要继续读取上次的记录吗？"
        )
        if answer:
            results = load_progress()
            print(f"已恢复 {len(results)} 条记录。")
        else:
            clear_progress()

    while True:
        panel.set_mode("未确定")
        panel.set_step("请选择编号")
        panel.clear_temp()
        panel.clear_status()
        
        current_data = {
            "编号": "",
            "长": "",
            "宽": ""
        }
        small_face_candidates = []

        print("\n=== 新的一组 ===")
        print("在 CAD 中依次点选：编号 → 分流（编号/长标注） → 宽标注")
        print("如需结束，请在任一步按 Esc 或取消选择。")

        try:
            if pending_code:
                code1 = pending_code
                pending_code = None
                panel.set_step("已读取缓存编号")
                panel.set_status(f"已自动带入下一组编号：{code1}")
            else:
                code1, err = pick_code_with_retry(
                    doc,
                    panel,
                    step_text="请选择编号",
                    prompt="请选择编号"
                )
                if err == "cancel":
                    break
            
            current_data["编号"] = code1
            panel.set_temp(code=code1)
            
            second_obj, second_type, err = pick_second_entity_with_retry(doc, panel)
            if err == "cancel":
                break

            second_text = None
            if second_type == "text":
                second_text, err_msg = read_code_entity(second_obj)
                if second_text is None:
                    print(err_msg)
                    panel.set_status(err_msg)
                    continue
                
            if second_type == "text" and is_code_text(second_text):
                panel.set_mode("范围模式")

                code2 = second_text

                merged_code = merge_code_range(code1, code2)
                if not merged_code:
                    msg = f"编号无法合并：{code1} × {code2}，本组跳过。"
                    print(msg)
                    panel.set_status(msg)
                    continue

                panel.clear_status()
                current_data["编号"] = merged_code
                panel.set_temp(code=merged_code)

                length_val, err = pick_dimension_with_retry(
                    doc,
                    panel,
                    step_text="请选择长尺寸",
                    prompt="请选择长尺寸（标注或文字）"
                )
                if err == "cancel":
                    break

                length_val = float(length_val)
                current_data["长"] = length_val
                panel.set_temp(length=cad_round_1(length_val))

                width_val, err = pick_dimension_with_retry(
                    doc,
                    panel,
                    step_text="请选择宽尺寸",
                    prompt="请选择宽尺寸（标注或文字）"
                )
                if err == "cancel":
                    break

                width_val = float(width_val)
                current_data["宽"] = width_val
                panel.set_temp(width=cad_round_1(width_val))
                panel.clear_status()

                row = {
                    "编号": merged_code,
                    "长": length_val,
                    "宽": width_val,
                    "长边1加工": "",
                    "长边2加工": "",
                    "短边1加工": "",
                    "短边2加工": "",
                    "左上角加工": "",
                    "右上角加工": "",
                    "左下角加工": "",
                    "右下角加工": ""
                }

                end_after_current = False

                while True:
                    next_code = process_post_width(doc, panel, small_face_candidates)

                    if next_code == "__END_AFTER_CURRENT__":
                        end_after_current = True
                        next_code = None

                    if next_code:
                        pending_code = next_code

                    print("本组 small_face_candidates =", small_face_candidates)

                    assign_result = assign_small_faces_to_edges(small_face_candidates, panel.root)

                    if assign_result["action"] == "back":
                        panel.set_status("返回继续补选小面加工。")
                        continue

                    if assign_result["action"] == "cancel":
                        msg = "用户取消了小面加工归类，本组小面加工保持空值。"
                        print(msg)
                        panel.set_status(msg)
                        break

                    if assign_result["action"] == "confirm":
                        assignments = assign_result["data"]
                        apply_small_face_assignments_to_row(row, assignments)
                        print("本组小面加工归类结果 =", assignments)
                        break
                    
                results.append(row)
                save_progress(results)

                panel.add_record({
                    "编号": row["编号"],
                    "长": cad_round_0(row["长"]),
                    "宽": cad_round_0(row["宽"]),
                    "长边1加工": row.get("长边1加工", ""),
                    "长边2加工": row.get("长边2加工", ""),
                    "短边1加工": row.get("短边1加工", ""),
                    "短边2加工": row.get("短边2加工", ""),
                    "左上角加工": row.get("左上角加工", ""),
                    "右上角加工": row.get("右上角加工", ""),
                    "左下角加工": row.get("左下角加工", ""),
                    "右下角加工": row.get("右下角加工", ""),
                })
                print("已记录（范围）：", row)
                
                if end_after_current:
                    print("当前组已保存，结束整个录入流程。")
                    break

            elif second_type == "dim" or (second_type == "text" and not is_code_text(second_text)):
                panel.set_mode("单编号模式")

                length_val, err_msg = read_size_entity(second_obj)
                if length_val is None:
                    print(err_msg)
                    panel.set_status(err_msg)
                    continue

                panel.clear_status()
                length_val = float(length_val)
                current_data["长"] = length_val
                panel.set_temp(length=cad_round_1(length_val))

                width_val, err = pick_dimension_with_retry(
                    doc,
                    panel,
                    step_text="请选择宽尺寸",
                    prompt="请选择宽尺寸（标注或文字）"
                )
                if err == "cancel":
                    break

                width_val = float(width_val)
                current_data["宽"] = width_val
                panel.set_temp(width=cad_round_1(width_val))
                panel.clear_status()

                row = {
                    "编号": code1,
                    "长": length_val,
                    "宽": width_val,
                    "长边1加工": "",
                    "长边2加工": "",
                    "短边1加工": "",
                    "短边2加工": "",
                    "左上角加工": "",
                    "右上角加工": "",
                    "左下角加工": "",
                    "右下角加工": ""
                }

                end_after_current = False

                while True:
                    next_code = process_post_width(doc, panel, small_face_candidates)

                    if next_code == "__END_AFTER_CURRENT__":
                        end_after_current = True
                        next_code = None

                    if next_code:
                        pending_code = next_code

                    print("本组 small_face_candidates =", small_face_candidates)

                    assign_result = assign_small_faces_to_edges(small_face_candidates, panel.root)

                    if assign_result["action"] == "back":
                        panel.set_status("返回继续补选小面加工。")
                        continue

                    if assign_result["action"] == "cancel":
                        msg = "用户取消了小面加工归类，本组小面加工保持空值。"
                        print(msg)
                        panel.set_status(msg)
                        break

                    if assign_result["action"] == "confirm":
                        assignments = assign_result["data"]
                        apply_small_face_assignments_to_row(row, assignments)
                        print("本组小面加工归类结果 =", assignments)
                        break

                results.append(row)
                save_progress(results)

                panel.add_record({
                    "编号": row["编号"],
                    "长": cad_round_0(row["长"]),
                    "宽": cad_round_0(row["宽"]),
                    "长边1加工": row.get("长边1加工", ""),
                    "长边2加工": row.get("长边2加工", ""),
                    "短边1加工": row.get("短边1加工", ""),
                    "短边2加工": row.get("短边2加工", ""),
                    "左上角加工": row.get("左上角加工", ""),
                    "右上角加工": row.get("右上角加工", ""),
                    "左下角加工": row.get("左下角加工", ""),
                    "右下角加工": row.get("右下角加工", ""),
                })
                print("已记录：", row)
                
                if end_after_current:
                    print("当前组已保存，结束整个录入流程。")
                    break

            else:
                _, obj_name = get_entity_type_with_name(second_obj)
                msg = f"第二个选择的对象类型无效，你选的是：{obj_name}"
                print(msg)
                panel.set_status(msg)
                continue

        except DiscardCurrentGroup:
            print("已放弃当前组，回到下一组。")
            panel.set_status("已放弃当前组，请重新录入。")
            panel.clear_temp()
            continue

        except EndWholeInput:
            print("用户结束整个录入流程。")
            break

    if results:
        process_names = collect_small_face_types(results)
        print("本次搜索到的小面加工类型：", process_names)

        config_path = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            "symbol_config.json"
        )
        print("config_path =", config_path)
        config = load_config(config_path)

        if not process_names:
            print("本次没有需要映射的小面加工类型，直接导出。")
            process_symbol_map = {}

            export_results_to_production_sheet(
                results,
                process_symbol_map,
                config,
                panel.root
            )

        else:
            symbol_options = get_symbol_options_from_config(config)
            print("symbol_options =", symbol_options)

            process_symbol_map = ask_small_face_symbol_mapping(
                process_names,
                symbol_options,
                panel.root
            )

            if process_symbol_map is None:
                print("用户取消图示选择，本次继续导出，但不绘制小面加工图示。")
                process_symbol_map = {}

            else:
                print("process_symbol_map =", process_symbol_map)

                process_symbol_map = {
                    k: v for k, v in process_symbol_map.items()
                    if v and v != "不绘制"
                }

            export_results_to_production_sheet(
                results,
                process_symbol_map,
                config,
                panel.root
            )

    else:
        print("没有可导出的数据。")
        
if __name__ == "__main__":
    main()
